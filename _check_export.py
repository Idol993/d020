import json, csv

with open('storage/exports/rb.json', encoding='utf-8') as f:
    jd = json.load(f)
print(f'JSON:  {len(jd)} records, {len(jd[0].keys())} fields')
jf = list(jd[0].keys())
print(f'  Fields: {jf}')

with open('storage/exports/rb.csv', encoding='utf-8-sig') as f:
    rows = list(csv.DictReader(f))
print(f'CSV:   {len(rows)} records, {len(rows[0].keys())} fields')
cf = list(rows[0].keys())
print(f'  Fields: {cf}')

from openpyxl import load_workbook
wb = load_workbook('storage/exports/rb.xlsx')
ws = wb.active
xf = [c.value for c in ws[1]]
print(f'EXCEL: {ws.max_row-1} records, {len(xf)} fields')
print(f'  Fields: {xf}')

print()
counts = {len(jd), len(rows), ws.max_row-1}
if len(counts) == 1:
    print(f'COUNT MATCH: {len(jd)} records')
else:
    print(f'COUNT MISMATCH: json={len(jd)} csv={len(rows)} xl={ws.max_row-1}')

if jf == cf == xf:
    print('FIELDS MATCH (all 15 fields)')
else:
    print('FIELDS MISMATCH')
    if set(jf) != set(cf):
        print(f'  json-csv: {set(jf).symmetric_difference(set(cf))}')
    if set(jf) != set(xf):
        print(f'  json-xl: {set(jf).symmetric_difference(set(xf))}')

print()
print('Record 0 (manual rollback):')
r = jd[0]
print(f'  event_id={r["event_id"]}, trigger_type={r["trigger_type"]}, trigger_metric={r["trigger_metric"]}')
print(f'  trigger_value={r["trigger_value"]}, affected_station_count={r["affected_station_count"]}')
print(f'  rollback_successful={r["rollback_successful"]}')
print()
print('Record 1 (auto circuit breaker):')
r = jd[1]
print(f'  event_id={r["event_id"]}, trigger_type={r["trigger_type"]}, trigger_metric={r["trigger_metric"]}')
print(f'  trigger_value={r["trigger_value"]}, threshold={r["threshold"]}')
print(f'  affected_station_count={r["affected_station_count"]}, rollback_successful={r["rollback_successful"]}')
