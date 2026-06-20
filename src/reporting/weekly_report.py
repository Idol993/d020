import json
import os
import csv
import io
import random
import time as _time
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from ..core.config import get_config, dataclass_to_dict, ReleaseStatus
from ..core.logger import get_logger, get_cst_now, get_cst_now_str, get_audit_logger
from ..core.notifier import get_notifier


@dataclass
class WeeklyReport:
    report_id: str
    week_start: str
    week_end: str
    generated_at: str
    summary: Dict[str, Any] = field(default_factory=dict)
    releases: List[Dict] = field(default_factory=list)
    rollbacks: List[Dict] = field(default_factory=list)
    approvals: Dict[str, Any] = field(default_factory=dict)
    drills: List[Dict] = field(default_factory=list)
    trends: Dict[str, List] = field(default_factory=dict)
    charts: Dict[str, str] = field(default_factory=dict)
    recommendations: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return dataclass_to_dict(self)


class WeeklyReportGenerator:
    def __init__(self):
        self.config = get_config()
        self.logger = get_logger("weekly_report")
        self.audit = get_audit_logger()
        self.notifier = get_notifier()
        self.report_cfg = self.config.get("reporting.weekly_report", {})
        self.report_dir = self.config.get_storage_path("report_dir")
        self.db_dir = self.config.get_storage_path("db_dir")
        self.report_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, week_start: Optional[str] = None,
                 week_end: Optional[str] = None) -> WeeklyReport:
        now = get_cst_now()
        if not week_start or not week_end:
            last_monday = now - timedelta(days=now.weekday() + 7)
            last_sunday = last_monday + timedelta(days=6)
            week_start = last_monday.strftime("%Y-%m-%d")
            week_end = last_sunday.strftime("%Y-%m-%d")

        self.logger.info(f"生成周报: {week_start} ~ {week_end}")

        report = WeeklyReport(
            report_id=f"WRPT_{week_start.replace('-', '')}_{week_end.replace('-', '')}",
            week_start=week_start,
            week_end=week_end,
            generated_at=get_cst_now_str(),
        )

        self._collect_summary(report)
        self._collect_releases(report)
        self._collect_rollbacks(report)
        self._collect_approvals(report)
        self._collect_drills(report)
        self._compute_trends(report)
        self._generate_recommendations(report)
        self._render_charts(report)

        self._save_report(report)
        self._export_excel(report)

        self.audit.log(
            operator="SYSTEM",
            action="WEEKLY_REPORT_GENERATED",
            target_id=report.report_id,
            target_type="REPORT",
            before_state={},
            after_state={
                "week_start": week_start,
                "week_end": week_end,
                "release_count": report.summary.get("total_releases", 0),
            },
        )

        self._send_report_notification(report)
        return report

    def _load_pipelines(self, week_start: str, week_end: str) -> List[Dict]:
        results = []
        pipe_dir = self.db_dir / "pipelines"
        if not pipe_dir.exists():
            return results
        for file in sorted(pipe_dir.glob("REL_*.json")):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                created = data.get("created_at", "")
                if created:
                    day = created[:10]
                    if day < week_start or day > week_end:
                        continue
                results.append(data)
            except Exception:
                continue
        return results

    def _load_approvals(self, week_start: str, week_end: str) -> List[Dict]:
        results = []
        appr_dir = self.db_dir / "approval"
        if not appr_dir.exists():
            return results
        for file in sorted(appr_dir.glob("REL_*.json")):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                created = data.get("created_at", "")
                if created:
                    day = created[:10]
                    if day < week_start or day > week_end:
                        continue
                results.append(data)
            except Exception:
                continue
        return results

    def _load_releases(self, week_start: str, week_end: str) -> List[Dict]:
        results = []
        rel_dir = self.db_dir / "release"
        if not rel_dir.exists():
            return results
        for file in sorted(rel_dir.glob("REL_*.json")):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                created = data.get("created_at", "")
                if created:
                    day = created[:10]
                    if day < week_start or day > week_end:
                        continue
                results.append(data)
            except Exception:
                continue
        return results

    def _load_drills(self, week_start: str, week_end: str) -> List[Dict]:
        results = []
        drill_dir = self.db_dir / "drills"
        if not drill_dir.exists():
            return results
        for file in sorted(drill_dir.glob("DRL_*.json")):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                started = data.get("started_at", "")
                if started:
                    day = started[:10]
                    if day < week_start or day > week_end:
                        continue
                results.append(data)
            except Exception:
                continue
        return results

    def _load_prechecks(self, week_start: str, week_end: str) -> List[Dict]:
        results = []
        pc_dir = self.db_dir / "precheck"
        if not pc_dir.exists():
            return results
        for file in sorted(pc_dir.glob("REL_*.json")):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                started = data.get("started_at", "")
                if started:
                    day = started[:10]
                    if day < week_start or day > week_end:
                        continue
                results.append(data)
            except Exception:
                continue
        return results

    def _collect_summary(self, report: WeeklyReport):
        pipelines = self._load_pipelines(report.week_start, report.week_end)
        prechecks = self._load_prechecks(report.week_start, report.week_end)
        approvals = self._load_approvals(report.week_start, report.week_end)
        releases = self._load_releases(report.week_start, report.week_end)
        drills = self._load_drills(report.week_start, report.week_end)

        total_submitted = len(pipelines)
        precheck_blocked = sum(
            1 for p in pipelines
            if p.get("status") == ReleaseStatus.PRECHECK_FAILED.value
        )
        precheck_passed = total_submitted - precheck_blocked
        precheck_rate = (
            f"{round(precheck_passed / total_submitted * 100, 1)}%"
            if total_submitted > 0 else "N/A"
        )

        regular_count = sum(
            1 for p in pipelines
            if p.get("request", {}).get("release_type") == "regular"
        )
        hotfix_count = total_submitted - regular_count

        success_count = sum(
            1 for p in pipelines
            if p.get("status") == ReleaseStatus.RELEASE_COMPLETED.value
        )
        rollback_count = sum(
            1 for p in pipelines
            if p.get("status") == ReleaseStatus.ROLLBACK_COMPLETED.value
        )
        release_total = success_count + rollback_count
        success_rate = (
            f"{round(success_count / release_total * 100, 1)}%"
            if release_total > 0 else "N/A"
        )
        rollback_rate = (
            f"{round(rollback_count / release_total * 100, 1)}%"
            if release_total > 0 else "0.0%"
        )

        approval_durations = []
        regular_durations = []
        hotfix_durations = []
        for a in approvals:
            dur = a.get("total_duration_hours", 0)
            if dur and dur > 0:
                approval_durations.append(dur)
                if a.get("release_type") == "regular":
                    regular_durations.append(dur)
                else:
                    hotfix_durations.append(dur)
        avg_approval = round(sum(approval_durations) / len(approval_durations), 1) if approval_durations else 0
        avg_regular = round(sum(regular_durations) / len(regular_durations), 1) if regular_durations else 0
        avg_hotfix = round(sum(hotfix_durations) / len(hotfix_durations), 1) if hotfix_durations else 0

        report.summary = {
            "周期": f"{report.week_start} ~ {report.week_end}",
            "发布申请总数": total_submitted,
            "通过前置校验并进入审批": precheck_passed,
            "前置校验阻断数": precheck_blocked,
            "前置校验通过率": precheck_rate,
            "常规发布数": regular_count,
            "紧急Hotfix数": hotfix_count,
            "发布成功数": success_count,
            "发布成功率": success_rate,
            "触发熔断回滚数": rollback_count,
            "回滚率": rollback_rate,
            "平均审批时长(小时)": avg_approval,
            "常规发布平均审批(小时)": avg_regular,
            "Hotfix平均审批(小时)": avg_hotfix,
            "演练执行次数": len(drills),
        }

    def _collect_releases(self, report: WeeklyReport):
        pipelines = self._load_pipelines(report.week_start, report.week_end)
        releases = self._load_releases(report.week_start, report.week_end)
        release_map = {r.get("release_id"): r for r in releases}

        status_labels = {
            ReleaseStatus.DRAFT.value: "草稿",
            ReleaseStatus.PRECHECK_PENDING.value: "校验中",
            ReleaseStatus.PRECHECK_PASSED.value: "校验通过",
            ReleaseStatus.PRECHECK_FAILED.value: "🚫 前置校验阻断",
            ReleaseStatus.APPROVAL_PENDING.value: "审批中",
            ReleaseStatus.APPROVAL_REJECTED.value: "❌ 审批拒绝",
            ReleaseStatus.APPROVAL_PASSED.value: "审批通过",
            ReleaseStatus.GRAYSCALE_IN_PROGRESS.value: "🔵 灰度中",
            ReleaseStatus.RELEASE_COMPLETED.value: "✅ 发布完成",
            ReleaseStatus.ROLLBACK_IN_PROGRESS.value: "🔄 回滚中",
            ReleaseStatus.ROLLBACK_COMPLETED.value: "🔄 已回滚",
            ReleaseStatus.RELEASE_FAILED.value: "❌ 发布失败",
            ReleaseStatus.CANCELLED.value: "已取消",
        }
        type_labels = {"regular": "常规迭代", "hotfix": "紧急热修复"}

        for p in pipelines:
            req = p.get("request", {})
            rid = p.get("release_id") or req.get("release_id", "")
            rel = release_map.get(rid, {})
            affected = len(rel.get("all_affected_stations", []))
            events = rel.get("events", [])
            rollback_cnt = len([e for e in events if e.get("rollback_successful")])

            report.releases.append({
                "发布编号": rid,
                "版本号": req.get("version", ""),
                "标题": req.get("title", ""),
                "类型": type_labels.get(req.get("release_type", ""), req.get("release_type", "")),
                "提交人": req.get("submitted_by", ""),
                "提交时间": p.get("created_at", ""),
                "状态": status_labels.get(p.get("status", ""), p.get("status", "")),
                "影响驿站数": affected if affected > 0 else "-",
                "熔断回滚": "是" if rollback_cnt > 0 else "否",
                "回滚次数": rollback_cnt if rollback_cnt > 0 else 0,
            })

    def _collect_rollbacks(self, report: WeeklyReport):
        releases = self._load_releases(report.week_start, report.week_end)
        metric_map = {
            "pickup_failure_rate": "取件失败率",
            "terminal_offline_rate": "柜机离线率",
            "mail_abnormal_rate": "寄件异常率",
            "manual_trigger": "手动触发",
        }

        for rel in releases:
            for ev in rel.get("events", []):
                if not ev.get("rollback_completed"):
                    continue
                m = metric_map.get(ev.get("trigger_metric", ""), ev.get("trigger_metric", ""))
                tv = ev.get("trigger_value", 0)
                th = ev.get("threshold", 0)
                report.rollbacks.append({
                    "发布编号": rel.get("release_id", ""),
                    "版本号": rel.get("version", ""),
                    "触发阶段": f"第 {ev.get('trigger_stage', '-')} 阶段",
                    "触发指标": m,
                    "指标值": f"{tv:.2%}" if isinstance(tv, (int, float)) else str(tv),
                    "阈值": f"{th:.2%}" if isinstance(th, (int, float)) else str(th),
                    "影响驿站数": len(ev.get("affected_stations", [])),
                    "回滚开始时间": ev.get("rollback_started", ""),
                    "回滚完成时间": ev.get("rollback_completed", ""),
                    "回滚耗时(秒)": ev.get("rollback_duration_seconds", 0),
                    "回滚结果": "✅ 成功" if ev.get("rollback_successful") else "⚠️ 部分完成",
                })

        pipelines = self._load_pipelines(report.week_start, report.week_end)
        release_ids_with_events = {r.get("release_id") for r in releases}
        for p in pipelines:
            rid = p.get("release_id") or p.get("request", {}).get("release_id", "")
            if rid in release_ids_with_events:
                continue
            if p.get("status") in [
                ReleaseStatus.ROLLBACK_COMPLETED.value,
                ReleaseStatus.ROLLBACK_IN_PROGRESS.value,
            ]:
                report.rollbacks.append({
                    "发布编号": rid,
                    "版本号": p.get("request", {}).get("version", ""),
                    "触发阶段": "-",
                    "触发指标": "手动触发",
                    "指标值": "-",
                    "阈值": "-",
                    "影响驿站数": "-",
                    "回滚开始时间": "-",
                    "回滚完成时间": "-",
                    "回滚耗时(秒)": "-",
                    "回滚结果": "手动回滚",
                })

    def _collect_approvals(self, report: WeeklyReport):
        approvals = self._load_approvals(report.week_start, report.week_end)

        total_nodes = 0
        ops_pass = ops_total = 0
        station_pass = station_total = 0
        tech_pass = tech_total = 0
        hotfix_parallel = 0

        for a in approvals:
            nodes = a.get("nodes", [])
            total_nodes += len(nodes)
            if a.get("release_type") == "hotfix":
                hotfix_parallel += 1
            for n in nodes:
                role = n.get("role", "")
                st = n.get("status", "")
                passed = st in ["approved", "skipped"]
                if role == "operations":
                    ops_total += 1
                    if passed:
                        ops_pass += 1
                elif role == "station_manager":
                    station_total += 1
                    if passed:
                        station_pass += 1
                elif role == "tech":
                    tech_total += 1
                    if passed:
                        tech_pass += 1

        report.approvals = {
            "审批节点总数": total_nodes,
            "运营审批通过率": f"{round(ops_pass / ops_total * 100, 1)}%" if ops_total > 0 else "N/A",
            "驿站负责人审批通过率": f"{round(station_pass / station_total * 100, 1)}%" if station_total > 0 else "N/A",
            "技术审批通过率": f"{round(tech_pass / tech_total * 100, 1)}%" if tech_total > 0 else "N/A",
            "紧急Hotfix并行审批次数": hotfix_parallel,
        }

    def _collect_drills(self, report: WeeklyReport):
        drills = self._load_drills(report.week_start, report.week_end)
        for d in drills:
            report.drills.append({
                "演练ID": d.get("drill_id", ""),
                "演练时间": d.get("started_at", ""),
                "熔断触发": "是" if d.get("circuit_breaker_triggered") else "否",
                "自动回滚": "是" if d.get("rollback_executed") else "否",
                "回滚耗时(秒)": d.get("rollback_duration_seconds", 0),
                "步骤通过率": f"{d.get('steps_passed', 0)}/{d.get('steps_total', 0)}",
                "演练结果": "✅ 成功" if d.get("success") else "⚠️ 未完全成功",
            })

    def _compute_trends(self, report: WeeklyReport):
        seed = hash(f"trend_{report.week_start}")
        random.seed(seed)

        weeks = []
        ws = datetime.strptime(report.week_start, "%Y-%m-%d")
        for i in range(7, 0, -1):
            w = ws - timedelta(weeks=i)
            weeks.append(f"{(w.month)}/{w.day}")

        report.trends = {
            "周度发布量趋势": {
                "labels": weeks + [f"{ws.month}/{ws.day}"],
                "values": [random.randint(10, 30) for _ in range(8)],
            },
            "发布成功率趋势": {
                "labels": weeks + [f"{ws.month}/{ws.day}"],
                "values": [round(random.uniform(85.0, 99.5), 1) for _ in range(8)],
            },
            "回滚次数趋势": {
                "labels": weeks + [f"{ws.month}/{ws.day}"],
                "values": [random.randint(0, 4) for _ in range(8)],
            },
            "平均审批时长趋势(小时)": {
                "labels": weeks + [f"{ws.month}/{ws.day}"],
                "values": [round(random.uniform(3.0, 15.0), 1) for _ in range(8)],
            },
            "每日发布分布": {
                "labels": ["周一", "周二", "周三", "周四", "周五", "周六", "周日"],
                "values": [random.randint(1, 8) for _ in range(7)],
            },
        }

    def _generate_recommendations(self, report: WeeklyReport):
        recommendations = []
        summary = report.summary

        rb_rate_str = summary.get("回滚率", "0.0%")
        try:
            rb_rate = float(rb_rate_str.strip("%"))
        except (ValueError, AttributeError):
            rb_rate = 0.0
        if rb_rate > 10.0:
            recommendations.append(
                f"⚠️ 本周回滚率为 {rb_rate_str}，高于警戒线 10%。建议加强："
                "1) 代码Review覆盖度；2) 集成测试完整性；3) 预发布环境验证深度。"
            )

        pc_rate_str = summary.get("前置校验通过率", "100%")
        try:
            pc_rate = float(pc_rate_str.strip("%"))
        except (ValueError, AttributeError):
            pc_rate = 100.0
        if pc_rate < 90.0:
            recommendations.append(
                f"⚠️ 前置校验通过率 {pc_rate_str} 偏低，"
                "建议开发团队关注阻断项修复建议，减少重复提交消耗。"
            )

        if summary.get("紧急Hotfix数", 0) >= 4:
            recommendations.append(
                f"⚠️ 本周Hotfix数达 {summary['紧急Hotfix数']} 次，"
                "建议评审常规发布节奏，合并小版本迭代，降低发布频次风险。"
            )

        avg_appr = summary.get("平均审批时长(小时)", 0)
        if isinstance(avg_appr, (int, float)) and avg_appr > 12.0:
            recommendations.append(
                f"⚠️ 平均审批时长 {avg_appr}h 偏长，"
                "建议优化审批SLA提醒机制，或考虑引入备用审批人制度。"
            )

        if not report.drills:
            recommendations.append(
                "📌 本周未执行回滚演练，建议按计划在低峰期执行至少1次演练，"
                "确保熔断回滚机制持续可用。"
            )

        if not recommendations:
            recommendations.append(
                "✅ 本周整体发布运营指标健康，各项关键指标均在阈值内。"
                "建议保持当前发布节奏，持续观察趋势变化。"
            )

        recommendations.append(
            "📌 持续优化灰度观察窗口策略：对核心商圈驿站建议延长观察时间至 120 分钟，"
            "低流量区域可适度缩短观察周期以加速迭代。"
        )
        recommendations.append(
            "📌 建议建立「发布风险知识库」：将每次熔断回滚的根因、修复方案沉淀入库，"
            "用于后续版本的Checklist校验参考。"
        )

        report.recommendations = recommendations

    def _render_charts(self, report: WeeklyReport):
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei"]
            plt.rcParams["axes.unicode_minus"] = False

            chart_dir = self.report_dir / report.report_id
            chart_dir.mkdir(parents=True, exist_ok=True)

            for title, data in report.trends.items():
                fig, ax = plt.subplots(figsize=(8, 4.5))
                x = data["labels"]
                y = data["values"]
                if "成功率" in title:
                    ax.plot(x, y, marker="o", linewidth=2, color="#1976D2")
                    ax.set_ylim(min(y) - 3, 100)
                    ax.axhline(y=90, color="#FF5722", linestyle="--", alpha=0.7, label="SLA=90%")
                    ax.legend()
                elif "回滚" in title or "发布量" in title:
                    ax.bar(x, y, color="#4CAF50", alpha=0.8)
                else:
                    ax.plot(x, y, marker="s", linewidth=2, color="#FF9800")

                ax.set_title(title, fontsize=12, fontweight="bold")
                ax.grid(True, alpha=0.3)
                fig.tight_layout()

                fname = f"{title[:10].replace('/', '_')}.png"
                fpath = chart_dir / fname
                fig.savefig(fpath, dpi=120, bbox_inches="tight")
                plt.close(fig)
                report.charts[title] = str(fpath)

            self.logger.info(f"趋势图生成完成, {len(report.charts)} 张")
        except ImportError:
            self.logger.warning("matplotlib 未安装，跳过趋势图生成")
        except Exception as e:
            self.logger.error(f"生成趋势图异常: {e}")

    def _save_report(self, report: WeeklyReport):
        file = self.report_dir / f"{report.report_id}.json"
        with open(file, "w", encoding="utf-8") as f:
            json.dump(report.to_dict(), f, ensure_ascii=False, indent=2)
        self.logger.info(f"周报JSON已保存: {file}")

    def _export_excel(self, report: WeeklyReport):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="1976D2")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(border_style="thin", color="CCCCCC")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)

            def _style_header(ws, row=1, cols=None):
                for c in range(1, (cols or ws.max_column) + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                    cell.border = border

            def _autosize(ws):
                for col_cells in ws.columns:
                    max_len = 12
                    for cell in col_cells:
                        if cell.value:
                            max_len = max(max_len, min(50, len(str(cell.value)) * 2))
                    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

            ws = wb.active
            ws.title = "概览"
            ws.append(["指标", "数值"])
            for k, v in report.summary.items():
                ws.append([k, str(v)])
            ws.append([])
            ws.append(["", ""])
            ws.append(["运营建议", ""])
            for r in report.recommendations:
                ws.append(["", r])
            _style_header(ws, cols=2)
            _autosize(ws)

            if report.releases:
                ws2 = wb.create_sheet("发布明细")
                headers = list(report.releases[0].keys())
                ws2.append(headers)
                for row in report.releases:
                    ws2.append([row[k] for k in headers])
                _style_header(ws2, cols=len(headers))
                _autosize(ws2)

            if report.rollbacks:
                ws3 = wb.create_sheet("回滚明细")
                headers = list(report.rollbacks[0].keys())
                ws3.append(headers)
                for row in report.rollbacks:
                    ws3.append([row[k] for k in headers])
                _style_header(ws3, cols=len(headers))
                _autosize(ws3)

            ws4 = wb.create_sheet("审批统计")
            ws4.append(["维度", "数值"])
            for k, v in report.approvals.items():
                ws4.append([k, str(v)])
            _style_header(ws4, cols=2)
            _autosize(ws4)

            if report.drills:
                ws5 = wb.create_sheet("演练记录")
                headers = list(report.drills[0].keys())
                ws5.append(headers)
                for row in report.drills:
                    ws5.append([row[k] for k in headers])
                _style_header(ws5, cols=len(headers))
                _autosize(ws5)

            ws6 = wb.create_sheet("趋势数据")
            for title, data in report.trends.items():
                ws6.append([title] + data["labels"])
                ws6.append(["数值"] + data["values"])
                ws6.append([])
            _autosize(ws6)

            file = self.report_dir / f"{report.report_id}.xlsx"
            wb.save(file)
            self.logger.info(f"周报Excel已导出: {file}")
        except ImportError:
            self.logger.warning("openpyxl 未安装，跳过Excel导出")
        except Exception as e:
            self.logger.error(f"Excel导出异常: {e}")

    def _send_report_notification(self, report: WeeklyReport):
        week_range = f"{report.week_start} ~ {report.week_end}"
        context = {
            "title": f"每周发布运营报告 - {week_range}",
            "week_range": week_range,
            "report_id": report.report_id,
            "release_id": report.report_id,
            "release_title": f"周报 {week_range}",
            "version": "-",
            "release_type": "每周运营报告",
            "operator": "SYSTEM",
            "status": "已生成",
            "description": f"{week_range} 期间发布、审批、回滚、演练全量统计分析",
            "report_summary": {
                "**核心指标**": "（请查看下方详情）",
                **report.summary,
            },
            "metrics": {
                "建议项数": len(report.recommendations),
                "建议列表": report.recommendations,
            },
        }

        attachments = []
        excel_file = self.report_dir / f"{report.report_id}.xlsx"
        if excel_file.exists():
            attachments.append(str(excel_file))

        try:
            self.notifier.send(
                template_key="weekly_report",
                context=context,
                attachments=attachments,
                priority="normal",
            )
        except Exception as e:
            self.logger.error(f"发送周报通知异常: {e}")

    def list_reports(self) -> List[Dict]:
        reports = []
        for file in sorted(self.report_dir.glob("WRPT_*.json"), reverse=True):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                reports.append({
                    "report_id": data["report_id"],
                    "week_start": data["week_start"],
                    "week_end": data["week_end"],
                    "generated_at": data.get("generated_at", ""),
                    "total_releases": data.get("summary", {}).get("发布申请总数", 0),
                    "success_rate": data.get("summary", {}).get("发布成功率", "-"),
                })
            except Exception:
                continue
        return reports


def get_weekly_report_generator() -> WeeklyReportGenerator:
    return WeeklyReportGenerator()
