import json
import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from ..core.config import get_config
from ..core.logger import get_logger, get_audit_logger, CST


class AuditQueryEngine:
    def __init__(self):
        self.config = get_config()
        self.logger = get_logger("audit_query")
        self.audit = get_audit_logger()
        self.db_dir = self.config.get_storage_path("db_dir")

    def query_release_records(self,
                              start_time: Optional[str] = None,
                              end_time: Optional[str] = None,
                              station: Optional[str] = None,
                              version: Optional[str] = None,
                              operator: Optional[str] = None,
                              release_type: Optional[str] = None,
                              status: Optional[str] = None) -> List[Dict]:
        self.logger.info(
            f"查询发布记录: time=[{start_time} ~ {end_time}], "
            f"station={station}, version={version}, operator={operator}, "
            f"type={release_type}, status={status}"
        )

        pipeline_dir = self.db_dir / "pipelines"
        if not pipeline_dir.exists():
            self.logger.info("pipelines目录不存在，查询结果为空")
            return []

        pipelines = {}
        for file in sorted(pipeline_dir.glob("*.json")):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                rid = data.get("release_id", file.stem)
                pipelines[rid] = data
            except Exception as e:
                self.logger.warning(f"解析流水线文件失败 {file}: {e}")

        sessions = {}
        release_dir = self.db_dir / "release"
        if release_dir.exists():
            for file in sorted(release_dir.glob("*.json")):
                try:
                    with open(file, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    rid = data.get("release_id", file.stem)
                    sessions[rid] = data
                except Exception as e:
                    self.logger.warning(f"解析发布会话文件失败 {file}: {e}")

        results = []
        for rid, pdata in pipelines.items():
            request = pdata.get("request", {})
            record = {
                "release_id": rid,
                "version": request.get("version", ""),
                "title": request.get("title", ""),
                "release_type": request.get("release_type", "regular"),
                "submitted_by": request.get("submitted_by", ""),
                "submitted_at": request.get("submitted_at", ""),
                "rollback_version": request.get("rollback_version", ""),
                "pipeline_status": pdata.get("status", ""),
                "current_step": pdata.get("current_step", ""),
                "created_at": pdata.get("created_at", ""),
                "updated_at": pdata.get("updated_at", ""),
                "error_message": pdata.get("error_message", ""),
                "step_history": pdata.get("step_history", []),
                "grayscale_status": "",
                "current_stage": 0,
                "total_stages": 0,
                "affected_stations": 0,
                "affected_station_list": [],
                "rollback_count": 0,
                "circuit_breaker_state": "",
                "completed_at": "",
            }

            sdata = sessions.get(rid)
            if sdata:
                events = sdata.get("events", [])
                record["grayscale_status"] = sdata.get("status", "")
                record["current_stage"] = sdata.get("current_stage_index", -1) + 1
                record["total_stages"] = len(sdata.get("stages", []))
                record["affected_stations"] = len(sdata.get("all_affected_stations", []))
                record["affected_station_list"] = sdata.get("all_affected_stations", [])
                record["rollback_count"] = len([e for e in events if e.get("rollback_successful")])
                record["circuit_breaker_state"] = sdata.get("circuit_breaker_state", "")
                record["completed_at"] = sdata.get("completed_at", "")

            if not self._match_release_filter(
                record, start_time, end_time, station,
                version, operator, release_type, status
            ):
                continue

            results.append(self._summarize_release_record(record))

        self.logger.info(f"查询到 {len(results)} 条发布记录")
        return sorted(results, key=lambda x: x.get("created_at", ""), reverse=True)

    def query_rollback_records(self,
                               start_time: Optional[str] = None,
                               end_time: Optional[str] = None,
                               station: Optional[str] = None,
                               version: Optional[str] = None,
                               release_id: Optional[str] = None) -> List[Dict]:
        self.logger.info(
            f"查询回滚记录: time=[{start_time} ~ {end_time}], "
            f"station={station}, version={version}, release_id={release_id}"
        )

        results = []

        release_dir = self.db_dir / "release"
        if release_dir.exists():
            for file in sorted(release_dir.glob("*.json")):
                try:
                    with open(file, "r", encoding="utf-8") as f:
                        sdata = json.load(f)

                    s_rid = sdata.get("release_id", "")
                    if release_id and s_rid != release_id:
                        continue
                    if version and sdata.get("version") != version:
                        continue

                    events = sdata.get("events", [])
                    for ev in events:
                        triggered_at = ev.get("triggered_at", "")
                        if start_time and triggered_at < start_time:
                            continue
                        if end_time and triggered_at > end_time + " 23:59:59":
                            continue
                        if station and station not in ev.get("affected_stations", []):
                            continue

                        trigger_metric = ev.get("trigger_metric", "")
                        if trigger_metric == "manual_trigger":
                            trigger_type = "手动回滚"
                        else:
                            trigger_type = "自动熔断"

                        trigger_value = ev.get("trigger_value", 0)
                        threshold = ev.get("threshold", 0)
                        if isinstance(trigger_value, (int, float)):
                            trigger_value_str = f"{trigger_value:.2%}"
                        else:
                            trigger_value_str = str(trigger_value)
                        if isinstance(threshold, (int, float)):
                            threshold_str = f"{threshold:.2%}"
                        else:
                            threshold_str = str(threshold)

                        results.append({
                            "event_id": ev.get("event_id"),
                            "release_id": s_rid,
                            "version": sdata.get("version"),
                            "rollback_version": sdata.get("rollback_version"),
                            "trigger_type": trigger_type,
                            "triggered_at": triggered_at,
                            "trigger_stage": ev.get("trigger_stage"),
                            "trigger_metric": trigger_metric,
                            "trigger_value": trigger_value_str,
                            "threshold": threshold_str,
                            "affected_station_count": len(ev.get("affected_stations", [])),
                            "rollback_started": ev.get("rollback_started"),
                            "rollback_completed": ev.get("rollback_completed"),
                            "rollback_duration_seconds": ev.get("rollback_duration_seconds"),
                            "rollback_successful": ev.get("rollback_successful"),
                        })
                except Exception:
                    continue

        pipeline_dir = self.db_dir / "pipelines"
        if pipeline_dir.exists():
            session_ids = set()
            if release_dir.exists():
                for f in release_dir.glob("*.json"):
                    session_ids.add(f.stem)

            for file in sorted(pipeline_dir.glob("*.json")):
                try:
                    with open(file, "r", encoding="utf-8") as f:
                        pdata = json.load(f)

                    rid = pdata.get("release_id", "")
                    if rid in session_ids:
                        continue

                    pstatus = pdata.get("status", "")
                    if "rollback" not in pstatus:
                        continue

                    if release_id and rid != release_id:
                        continue

                    request = pdata.get("request", {})
                    if version and request.get("version") != version:
                        continue

                    rollback_history = [
                        h for h in pdata.get("step_history", [])
                        if "ROLLBACK" in h.get("action", "")
                    ]
                    if not rollback_history:
                        continue

                    last_rb = rollback_history[-1]
                    rb_time = last_rb.get("timestamp", pdata.get("created_at", ""))
                    if start_time and rb_time < start_time:
                        continue
                    if end_time and rb_time > end_time + " 23:59:59":
                        continue

                    results.append({
                        "event_id": f"PL_{rid}",
                        "release_id": rid,
                        "version": request.get("version", ""),
                        "rollback_version": request.get("rollback_version", ""),
                        "trigger_type": "手动回滚",
                        "triggered_at": rb_time,
                        "trigger_stage": 0,
                        "trigger_metric": "manual_trigger",
                        "trigger_value": "N/A",
                        "threshold": "N/A",
                        "affected_station_count": 0,
                        "rollback_started": rb_time,
                        "rollback_completed": pdata.get("updated_at", ""),
                        "rollback_duration_seconds": "N/A",
                        "rollback_successful": pstatus == "rollback_completed",
                    })
                except Exception:
                    continue

        self.logger.info(f"查询到 {len(results)} 条回滚记录")
        return sorted(results, key=lambda x: x.get("triggered_at", ""), reverse=True)

    def query_drill_records(self, start_time: Optional[str] = None,
                            end_time: Optional[str] = None,
                            status: Optional[str] = None) -> List[Dict]:
        results = []
        drill_dir = self.db_dir / "drills"
        if drill_dir.exists():
            for file in sorted(drill_dir.glob("DRL_*.json"), reverse=True):
                try:
                    with open(file, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    if start_time and data.get("started_at", "") < start_time:
                        continue
                    if end_time and data.get("started_at", "") > end_time + " 23:59:59":
                        continue
                    if status and data.get("status") != status:
                        continue
                    results.append(data)
                except Exception:
                    continue
        return results

    def query_approval_records(self, start_time: Optional[str] = None,
                               end_time: Optional[str] = None,
                               release_id: Optional[str] = None,
                               approver: Optional[str] = None) -> List[Dict]:
        results = []
        approval_dir = self.db_dir / "approval"
        if approval_dir.exists():
            for file in sorted(approval_dir.glob("REL_*.json")):
                try:
                    with open(file, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    if start_time and data.get("created_at", "") < start_time:
                        continue
                    if end_time and data.get("created_at", "") > end_time + " 23:59:59":
                        continue
                    if release_id and data.get("release_id") != release_id:
                        continue
                    if approver:
                        matched = any(
                            approver in n.get("approvers", []) or n.get("approved_by") == approver
                            for n in data.get("nodes", [])
                        )
                        if not matched:
                            continue

                    node_details = []
                    for n in data.get("nodes", []):
                        node_details.append({
                            "role": n.get("role"),
                            "name": n.get("name"),
                            "status": n.get("status"),
                            "approved_by": n.get("approved_by"),
                            "approved_at": n.get("approved_at"),
                            "comment": n.get("comment"),
                        })

                    results.append({
                        "release_id": data.get("release_id"),
                        "release_type": data.get("release_type"),
                        "channel_name": data.get("channel_name"),
                        "created_at": data.get("created_at"),
                        "started_at": data.get("started_at"),
                        "completed_at": data.get("completed_at"),
                        "status": data.get("status"),
                        "overall_passed": data.get("overall_passed"),
                        "total_duration_hours": data.get("total_duration_hours"),
                        "rejected_by": data.get("rejected_by"),
                        "rejected_reason": data.get("rejected_reason"),
                        "nodes": node_details,
                    })
                except Exception:
                    continue
        return sorted(results, key=lambda x: x["created_at"], reverse=True)

    def _match_release_filter(self, record: Dict, start_time, end_time,
                              station, version, operator, release_type,
                              status) -> bool:
        created = record.get("created_at", "")
        if start_time and created < start_time:
            return False
        if end_time and created > end_time + " 23:59:59":
            return False
        if version and record.get("version") != version:
            return False
        if operator and record.get("submitted_by") != operator:
            return False
        if release_type and record.get("release_type") != release_type:
            return False
        if station:
            affected = record.get("affected_station_list", [])
            if station not in affected:
                return False
        if status:
            pstatus = record.get("pipeline_status", "")
            gs = record.get("grayscale_status", "")
            if status == "rollback":
                if "rollback" not in pstatus and "rollback" not in gs and gs != "circuit_breaker_triggered":
                    return False
            elif status == "success":
                if pstatus != "release_completed" and gs != "completed":
                    return False
            elif status == "in_progress":
                if "in_progress" not in pstatus and gs not in ("in_progress", "created"):
                    return False
        return True

    def _summarize_release_record(self, record: Dict) -> Dict:
        pstatus = record.get("pipeline_status", "")
        gs = record.get("grayscale_status", "")

        pstatus_map = {
            "draft": "草稿",
            "precheck_pending": "校验中",
            "precheck_passed": "校验通过",
            "precheck_failed": "校验失败",
            "approval_pending": "审批中",
            "approval_passed": "审批通过",
            "approval_rejected": "审批拒绝",
            "grayscale_in_progress": "灰度发布中",
            "release_completed": "发布完成",
            "rollback_in_progress": "回滚中",
            "rollback_completed": "已回滚",
            "release_failed": "发布失败",
            "cancelled": "已取消",
        }

        gs_map = {
            "created": "灰度创建",
            "in_progress": "灰度进行中",
            "completed": "发布成功",
            "rollback_completed": "已回滚",
            "rollback_failed": "回滚失败",
            "circuit_breaker_triggered": "熔断触发",
            "cancelled": "已取消",
            "failed": "发布失败",
        }

        if gs:
            display_status = gs_map.get(gs, gs)
        else:
            display_status = pstatus_map.get(pstatus, pstatus)

        return {
            "release_id": record.get("release_id"),
            "version": record.get("version"),
            "title": record.get("title", ""),
            "release_type": record.get("release_type", "regular"),
            "submitted_by": record.get("submitted_by", ""),
            "rollback_version": record.get("rollback_version", ""),
            "status": display_status,
            "pipeline_status": pstatus,
            "grayscale_status": gs_map.get(gs, gs) if gs else "",
            "current_stage": record.get("current_stage", 0),
            "total_stages": record.get("total_stages", 0),
            "affected_stations": record.get("affected_stations", 0),
            "rollback_count": record.get("rollback_count", 0),
            "circuit_breaker_state": record.get("circuit_breaker_state", ""),
            "created_at": record.get("created_at", ""),
            "updated_at": record.get("updated_at", ""),
            "completed_at": record.get("completed_at", ""),
        }

    def export(self, records: List[Dict], output_path: str,
               format: str = "excel") -> str:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        if format == "csv":
            if records:
                keys = list(records[0].keys())
                with open(output, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
                    writer.writeheader()
                    for r in records:
                        flat = {}
                        for k, v in r.items():
                            if isinstance(v, (dict, list)):
                                flat[k] = json.dumps(v, ensure_ascii=False)
                            else:
                                flat[k] = v
                        writer.writerow(flat)
        elif format == "json":
            with open(output, "w", encoding="utf-8") as f:
                json.dump(records, f, ensure_ascii=False, indent=2, default=str)
        elif format == "excel":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                wb = Workbook()
                ws = wb.active
                if records:
                    keys = list(records[0].keys())
                    ws.append(keys)
                    hf = Font(bold=True, color="FFFFFF")
                    hfill = PatternFill("solid", fgColor="1976D2")
                    for c in range(1, len(keys) + 1):
                        cell = ws.cell(row=1, column=c)
                        cell.font = hf
                        cell.fill = hfill
                        cell.alignment = Alignment(horizontal="center")
                    for r in records:
                        row_data = []
                        for k in keys:
                            v = r.get(k, "")
                            if isinstance(v, (dict, list)):
                                v = json.dumps(v, ensure_ascii=False)
                            row_data.append(v)
                        ws.append(row_data)
                    for col in ws.columns:
                        mx = 12
                        for cell in col:
                            if cell.value:
                                mx = max(mx, min(50, len(str(cell.value)) * 2))
                        ws.column_dimensions[col[0].column_letter].width = mx
                wb.save(output)
            except ImportError:
                self.logger.warning("openpyxl未安装，降级为CSV导出")
                return self.export(records, str(output.with_suffix(".csv")), "csv")

        self.logger.info(f"记录已导出到 {output}, 共 {len(records)} 条")
        return str(output)


def get_audit_query_engine() -> AuditQueryEngine:
    return AuditQueryEngine()
